import openpyxl

wb = openpyxl.load_workbook('public\\QuestionsFormat.xlsx')
ws = wb.active

print("Template Column Structure:")
print("=" * 80)

headers = []
for i, cell in enumerate(ws[1], 1):
    value = str(cell.value or '')
    headers.append(value)
    print(f"Column {i}: {value}")

print("\n" + "=" * 80)
print(f"Total columns: {len(headers)}")
print("\nSample data (Row 2):")
for i, cell in enumerate(ws[2], 1):
    print(f"  Col {i}: {cell.value}")
